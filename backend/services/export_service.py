"""
渊博579 HR V7 — Export Service

Uses openpyxl to generate Excel workbooks.
All amounts in EUR. Dates formatted as YYYY-MM-DD.
"""
from __future__ import annotations
import io
from typing import List
from datetime import datetime


def _make_wb_title(ws, title: str, col_count: int) -> None:
    """Add a merged title cell at the top of a worksheet."""
    from openpyxl.styles import Font, Alignment, PatternFill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor="4F6EF7")
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    ws.row_dimensions[1].height = 28


def _add_header_row(ws, headers: list, row: int = 2) -> None:
    from openpyxl.styles import Font, Alignment, PatternFill
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill("solid", fgColor="E8EDF7")
    ws.row_dimensions[row].height = 20


def _auto_col_width(ws) -> None:
    """Auto-fit column widths based on content. Skips merged cells."""
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if isinstance(cell, MergedCell):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def export_employee_settlements_xlsx(rows: list, period: str) -> bytes:
    """Export employee settlement list to Excel."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"员工结算_{period}"

    headers = [
        "结算单号", "月份", "工号", "姓名", "职级", "仓库", "结算类型",
        "工时条数", "总工时(h)", "基础工资(€)", "班次补贴(€)", "扣款(€)",
        "税前工资(€)", "社保(€)", "综合成本(€)", "状态",
    ]
    _make_wb_title(ws, f"渊博579 员工月度结算 — {period}", len(headers))
    _add_header_row(ws, headers)

    for row_idx, r in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=r.settle_no)
        ws.cell(row=row_idx, column=2, value=r.period)
        ws.cell(row=row_idx, column=3, value=r.emp_no)
        ws.cell(row=row_idx, column=4, value=r.emp_name)
        ws.cell(row=row_idx, column=5, value=r.grade)
        ws.cell(row=row_idx, column=6, value=r.warehouse_code)
        ws.cell(row=row_idx, column=7, value=r.settlement_type)
        ws.cell(row=row_idx, column=8, value=r.timesheet_count)
        ws.cell(row=row_idx, column=9, value=round(r.total_hours, 2))
        ws.cell(row=row_idx, column=10, value=round(r.base_pay, 2))
        ws.cell(row=row_idx, column=11, value=round(r.bonus_pay, 2))
        ws.cell(row=row_idx, column=12, value=round(r.deduction, 2))
        ws.cell(row=row_idx, column=13, value=round(r.gross_pay, 2))
        ws.cell(row=row_idx, column=14, value=round(r.social_cost, 2))
        ws.cell(row=row_idx, column=15, value=round(r.total_cost, 2))
        ws.cell(row=row_idx, column=16, value=r.status)

    _auto_col_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_supplier_settlements_xlsx(rows: list, period: str) -> bytes:
    """Export supplier settlement list to Excel."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"供应商结算_{period}"

    headers = [
        "结算单号", "月份", "供应商", "业务线", "员工人数", "工时条数",
        "总工时(h)", "应付金额(€)", "发票号", "发票日期", "发票金额(€)", "状态",
    ]
    _make_wb_title(ws, f"渊博579 供应商月度结算 — {period}", len(headers))
    _add_header_row(ws, headers)

    for row_idx, r in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=r.settle_no)
        ws.cell(row=row_idx, column=2, value=r.period)
        ws.cell(row=row_idx, column=3, value=r.supplier_name)
        ws.cell(row=row_idx, column=4, value=r.biz_line)
        ws.cell(row=row_idx, column=5, value=r.employee_count)
        ws.cell(row=row_idx, column=6, value=r.timesheet_count)
        ws.cell(row=row_idx, column=7, value=round(r.total_hours, 2))
        ws.cell(row=row_idx, column=8, value=round(r.total_amount, 2))
        ws.cell(row=row_idx, column=9, value=r.invoice_no or "")
        ws.cell(row=row_idx, column=10, value=r.invoice_date.strftime('%Y-%m-%d') if r.invoice_date else "")
        ws.cell(row=row_idx, column=11, value=round(r.invoice_amount, 2))
        ws.cell(row=row_idx, column=12, value=r.status)

    _auto_col_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_project_settlements_xlsx(rows: list, period: str) -> bytes:
    """Export project gross profit analysis to Excel."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"项目毛利_{period}"

    headers = [
        "结算单号", "月份", "仓库", "业务线",
        "客户收入(€)", "自有人力成本(€)", "供应商成本(€)",
        "总成本(€)", "毛利(€)", "毛利率(%)", "状态",
    ]
    _make_wb_title(ws, f"渊博579 项目毛利分析 — {period}", len(headers))
    _add_header_row(ws, headers)

    for row_idx, r in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=r.settle_no)
        ws.cell(row=row_idx, column=2, value=r.period)
        ws.cell(row=row_idx, column=3, value=r.warehouse_code)
        ws.cell(row=row_idx, column=4, value=r.biz_line)
        ws.cell(row=row_idx, column=5, value=round(r.client_revenue, 2))
        ws.cell(row=row_idx, column=6, value=round(r.own_labor_cost, 2))
        ws.cell(row=row_idx, column=7, value=round(r.supplier_labor_cost, 2))
        ws.cell(row=row_idx, column=8, value=round(r.total_labor_cost, 2))
        ws.cell(row=row_idx, column=9, value=round(r.gross_profit, 2))
        ws.cell(row=row_idx, column=10, value=round(r.gross_margin * 100, 1))
        ws.cell(row=row_idx, column=11, value=r.status)

    _auto_col_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_commissions_xlsx(rows: list, monthly_map: dict, period: str) -> bytes:
    """Export commission agreements with monthly details to Excel.
    rows: list of CommissionRecord ORM objects
    monthly_map: {commission_id: [CommissionMonthly]} dict
    """
    from openpyxl import Workbook
    wb = Workbook()

    # Sheet 1: Commission agreements summary
    ws1 = wb.active
    ws1.title = f"返佣协议_{period}"
    headers1 = [
        "协议编号", "推荐人", "类型", "客户名称", "仓库",
        "层级", "返佣率(%)", "生效日期", "到期日期",
        "已付(€)", "待付(€)", "状态",
    ]
    _make_wb_title(ws1, f"渊博579 返佣协议汇总 — {period}", len(headers1))
    _add_header_row(ws1, headers1)
    for row_idx, r in enumerate(rows, start=3):
        ws1.cell(row=row_idx, column=1, value=r.commission_no)
        ws1.cell(row=row_idx, column=2, value=r.referrer_name)
        ws1.cell(row=row_idx, column=3, value="个人" if r.referrer_type == "individual" else "机构")
        ws1.cell(row=row_idx, column=4, value=r.client_name)
        ws1.cell(row=row_idx, column=5, value=r.client_warehouse or "")
        ws1.cell(row=row_idx, column=6, value=r.tier.upper())
        ws1.cell(row=row_idx, column=7, value=round(r.commission_rate, 1))
        ws1.cell(row=row_idx, column=8, value=r.validity_start.strftime('%Y-%m-%d') if r.validity_start else "")
        ws1.cell(row=row_idx, column=9, value=r.validity_end.strftime('%Y-%m-%d') if r.validity_end else "")
        ws1.cell(row=row_idx, column=10, value=round(r.total_paid, 2))
        ws1.cell(row=row_idx, column=11, value=round(r.total_pending, 2))
        ws1.cell(row=row_idx, column=12, value=r.status)
    _auto_col_width(ws1)

    # Sheet 2: Monthly details for the requested period
    ws2 = wb.create_sheet(title=f"月度明细_{period}")
    headers2 = [
        "协议编号", "推荐人", "客户名称", "月份",
        "发票额(€)", "返佣率(%)", "返佣额(€)", "付款状态", "付款日期",
    ]
    _make_wb_title(ws2, f"渊博579 返佣月度明细 — {period}", len(headers2))
    _add_header_row(ws2, headers2)
    row_idx = 3
    for r in rows:
        for m in monthly_map.get(r.id, []):
            ws2.cell(row=row_idx, column=1, value=r.commission_no)
            ws2.cell(row=row_idx, column=2, value=r.referrer_name)
            ws2.cell(row=row_idx, column=3, value=r.client_name)
            ws2.cell(row=row_idx, column=4, value=m.period)
            ws2.cell(row=row_idx, column=5, value=round(m.client_invoice_amount, 2))
            ws2.cell(row=row_idx, column=6, value=round(m.commission_rate, 1))
            ws2.cell(row=row_idx, column=7, value=round(m.commission_amount, 2))
            ws2.cell(row=row_idx, column=8, value=m.payment_status)
            ws2.cell(row=row_idx, column=9, value=m.payment_date.strftime('%Y-%m-%d') if m.payment_date else "")
            row_idx += 1
    _auto_col_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_quotation_xlsx(quotation, cost_calcs: list) -> bytes:
    """Export a quotation to German-format Excel.
    quotation: Quotation ORM object
    cost_calcs: list of CostCalculation ORM objects linked to this quotation
    """
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    wb = Workbook()

    # ── Sheet 1: Angebot ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Angebot"

    # Company header
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "渊博579 HR Dispatch Management GmbH — Angebot"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1E3A5F")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Quote meta
    ws["A3"] = "Angebot-Nr:"
    ws["B3"] = quotation.quote_no
    ws["A4"] = "Auftraggeber:"
    ws["B4"] = quotation.client_name
    ws["A5"] = "Kontakt:"
    ws["B5"] = quotation.client_contact or ""
    ws["A6"] = "Lager:"
    ws["B6"] = quotation.warehouse_code or ""
    ws["A7"] = "Projekttyp:"
    ws["B7"] = quotation.project_type or ""
    ws["A8"] = "Gültig bis:"
    ws["B8"] = quotation.valid_until.strftime("%d.%m.%Y") if quotation.valid_until else ""
    ws["A9"] = "Status:"
    ws["B9"] = quotation.status
    ws["D3"] = "Erstellt am:"
    ws["E3"] = quotation.created_at.strftime("%d.%m.%Y") if quotation.created_at else ""
    ws["D4"] = "Erstellt von:"
    ws["E4"] = quotation.created_by or ""
    ws["D5"] = "Genehmigt von:"
    ws["E5"] = quotation.approved_by or ""

    for row in range(3, 10):
        ws.cell(row=row, column=1).font = Font(bold=True, size=10)

    # Line items header
    row_start = 12
    ws.cell(row=row_start, column=1, value="Pos.")
    ws.cell(row=row_start, column=2, value="Leistungsart")
    ws.cell(row=row_start, column=3, value="Einheit")
    ws.cell(row=row_start, column=4, value="Menge")
    ws.cell(row=row_start, column=5, value="Einzelpreis (€)")
    ws.cell(row=row_start, column=6, value="Rabatt (%)")
    ws.cell(row=row_start, column=7, value="Nettobetrag (€)")
    for col in range(1, 8):
        cell = ws.cell(row=row_start, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="2E5FA3")
        cell.alignment = Alignment(horizontal="center")

    # Line items data
    items = []
    if quotation.items_json:
        try:
            items = json.loads(quotation.items_json)
        except Exception:
            items = []

    r = row_start + 1
    subtotal = 0.0
    for i, item in enumerate(items, start=1):
        ws.cell(row=r, column=1, value=i)
        label = item.get("label") or item.get("biz_line", "")
        ws.cell(row=r, column=2, value=f"{item.get('biz_line','')} {label}".strip())
        ws.cell(row=r, column=3, value=item.get("unit", ""))
        ws.cell(row=r, column=4, value=item.get("volume", 0))
        ws.cell(row=r, column=5, value=round(item.get("base_price", item.get("net_price", 0)), 4))
        ws.cell(row=r, column=6, value=f"{round(item.get('discount',0)*100, 0):.0f}%")
        ws.cell(row=r, column=7, value=round(item.get("amount", 0), 2))
        subtotal += item.get("amount", 0)
        r += 1

    # Hourly cost line if no items
    if not items and quotation.quote_hourly_rate:
        ws.cell(row=r, column=2, value="Personaldienstleistung (Stundenlohn)")
        ws.cell(row=r, column=3, value="Std.")
        ws.cell(row=r, column=5, value=round(quotation.quote_hourly_rate, 2))
        r += 1

    r += 1
    subtotal = round(subtotal, 2) or round((quotation.total_monthly_estimate or 0) / 1.19, 2)
    mwst = round(subtotal * 0.19, 2)
    total = round(subtotal + mwst, 2)

    ws.cell(row=r, column=6, value="Summe Netto:")
    ws.cell(row=r, column=7, value=subtotal)
    ws.cell(row=r, column=6).font = Font(bold=True)
    r += 1
    ws.cell(row=r, column=6, value="MwSt. 19%:")
    ws.cell(row=r, column=7, value=mwst)
    r += 1
    ws.cell(row=r, column=6, value="Gesamtbetrag Brutto:")
    ws.cell(row=r, column=7, value=total)
    ws.cell(row=r, column=6).font = Font(bold=True, size=12)
    ws.cell(row=r, column=7).font = Font(bold=True, size=12, color="1B5E20")

    # Cost summary if available
    if quotation.cost_total_per_hour:
        r += 2
        ws.cell(row=r, column=1, value="Kostenübersicht")
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        r += 1
        for label_de, val in [
            ("Bruttolohn/Std.", f"€{quotation.cost_hourly_rate:.4f}"),
            ("Sozialversicherung", f"{round((quotation.cost_social_rate or 0)*100, 0):.0f}%"),
            ("Verwaltungskosten", f"{round((quotation.cost_management_fee or 0)*100, 0):.0f}%"),
            ("Gesamtkosten/Std.", f"€{quotation.cost_total_per_hour:.4f}"),
            ("Angebotsrate/Std.", f"€{quotation.quote_hourly_rate:.2f}" if quotation.quote_hourly_rate else "-"),
            ("Ziel-Gewinnspanne", f"{round((quotation.quote_margin or 0)*100, 0):.0f}%"),
        ]:
            ws.cell(row=r, column=1, value=label_de)
            ws.cell(row=r, column=1).font = Font(bold=True, size=10)
            ws.cell(row=r, column=2, value=val)
            r += 1

    _auto_col_width(ws)

    # ── Sheet 2: Kostenanalyse ────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Kostenanalyse P1-P9")
    _make_wb_title(ws2, "Personalkosten nach Entgeltgruppe (P1-P9)", 5)
    _add_header_row(ws2, ["Entgeltgruppe", "Bruttolohn/Std. (€)", "Gesamtkosten/Std. (€)", "Empfohlen. Preis 20% (€)", "Empfohlen. Preis 25% (€)"])
    import backend.config as cfg
    from backend.services import cost_calculator
    for i, (grade, coeff) in enumerate(cfg.COEFFICIENTS.items(), start=3):
        r20 = cost_calculator.calc_full(grade=grade, target_margin=0.20)
        r25 = cost_calculator.calc_full(grade=grade, target_margin=0.25)
        ws2.cell(row=i, column=1, value=grade)
        ws2.cell(row=i, column=2, value=round(r20["gross_hourly"], 4))
        ws2.cell(row=i, column=3, value=round(r20["total_cost_per_hour"], 4))
        ws2.cell(row=i, column=4, value=round(r20["suggested_rate"], 2))
        ws2.cell(row=i, column=5, value=round(r25["suggested_rate"], 2))
    _auto_col_width(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
